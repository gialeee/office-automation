from openpyxl import load_workbook
import os


def read_file_size(filepath):
    return os.path.getsize(os.path.abspath(filepath))


def write_file_size(dir_path, excel_path):
    wb = load_workbook(excel_path)
    sheet = wb.active
    for row in range(2, sheet.max_row+1):
        if(sheet.cell(row, 11).value is not None):
            file_name = sheet.cell(row, 15).value
            sheet[f"L{row}"] = read_file_size(f"{dir_path}/{file_name}")

    wb.save('./file_migration_with_size.xlsx')

        
        
if __name__ == "__main__":
  excel_path = os.path.abspath("./file_migration.xlsx")
  dir_path = os.path.abspath("./file_migration")

  write_file_size(dir_path, excel_path)

  while True:
      try:
          user_input = input("*** 종료하려면 'q'를 입력해주세요.\n").lower()
          if user_input == 'q':
              break
      except:
          continue
  
