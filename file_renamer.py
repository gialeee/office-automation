from openpyxl import load_workbook
import os


filename_dict = {}

# 파일명과 파일 key값을 딕셔너리로 맵핑해주는 함수
def map_file2key(excel_path):
    try:
        wb = load_workbook(excel_path)
        sheet = wb.active
    except:
        print("\n[오류] 'file_migration.xlsx' 파일을 찾을 수 없습니다. 실행파일 경로를 다시 확인해주세요.\n")
    for row in range(2, sheet.max_row+1):
        if(sheet.cell(row, 11).value is not None):
            file_name = (sheet.cell(row, 11).value)
            file_key = (sheet.cell(row, 15).value)
            filename_dict[file_name] = file_key
        else:
            print("\n'file_migration.xlsx'에서 원본 파일명이 누락되었습니다. 지정된 위치에 값이 있는지 확인해주세요. \n(K : 칼럼명 / O : 파일 키)%n")

    return filename_dict


# 파일명을 수정하는 함수
def rename_file(dir_path, filename_dict):
    try:
        if os.path.isdir(dir_path):
            pass
    except:
        print("\n[오류] 'file_migration' 폴더를 찾을 수 없습니다. 실행파일 경로를 다시 확인해주세요.\n")

    for file_info in filename_dict.items():  # file_info[0] : FILE_NAME / file_info[1] : UPLOAD_FILE_NAME
        try:
            original = os.path.abspath(f"{dir_path}/{file_info[0]}")
            new =  os.path.abspath(f"{dir_path}/{file_info[1]}")

            print(f"ORIGINAL FILE NAME : {original} \n NEW FILE NAME : {new}\n")

            os.rename(original, new)
        except:
            print(f"\n *** [오류] '{file_info[0]}' 파일을 찾을 수 없습니다. 디렉토리 경로를 확인해주세요.\n")


if __name__ == "__main__":
  excel_path = os.path.abspath("./file_migration.xlsx")
  dir_path = os.path.abspath("./file_migration")
  filename_dict = map_file2key(excel_path)
  rename_file(dir_path, filename_dict)

  while True:
      try:
          user_input = input("*** 종료하려면 'q'를 입력해주세요.\n").lower()
          if user_input == 'q':
              break
      except:
          continue
